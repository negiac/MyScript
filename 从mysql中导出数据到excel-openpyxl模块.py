#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Date    : 2020-11-19 10:18:39
# @Author  : negiaC (lv-007@163.com)


import pymysql
#import xlwt
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import importlib
import sys
importlib.reload(sys)
import datetime
from email.mime.multipart import MIMEMultipart
from email.header import Header
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.application import MIMEApplication
import smtplib
from glob import glob
# import glob2 as glob  #python2.7
import sys
import os


class mysqlDB(object):
    """docstring for mysqlDB"""

    def __init__(self):
        self.host = 'IP地址
        self.port = port #int
        self.user = 'user'
        self.passwd = 'passwd'
        self.db='dbname'
        #self.db = ''
        self.charset = 'utf8'

    def conn_db(self):
        conn = pymysql.connect(host=self.host, port=self.port, user=self.user,
                               passwd=self.passwd, db=self.db, charset=self.charset)
        return conn

    # 使用cursor()方法获取操作游标
    def cursor_db(self):
        conn = self.conn_db()
        cursor = conn.cursor()
        return cursor


class excel(object):
    """docstring for excel"""
    # def __init__(self):
    #       self = self
    # 创建Excel

    def create_excel(self):
        #workbook = xlwt.Workbook(encoding='utf-8')
        workbook = Workbook()
        return workbook

    # 在Excel里创建第一张sheet表

    def save_data1_to_excel(self, cursor, workbook, sql, sheet_name):
        # 使用execute方法执行SQL语句,并将统计结果存储在effect_row变量中
        effect_row = cursor.execute(sql)
        # 打印总行数
        print(effect_row)
        # 获取所有的记录结果
        results = cursor.fetchall()
        # 获取上述SQL语句中的检索条件名称（将要成为Excel第一张sheet表的第一行的表头）
        fields = cursor.description
        print(fields)
        # 创建Excel中的一个sheet，并命名且为可重写状态
        #sheet = workbook.add_sheet(sheet_name, cell_overwrite_ok=True)
        sheet = workbook.create_sheet(title=sheet_name)

        # 写上字段信息 openpyxl 中行号/列号从1 开始
        for field in range(1, len(fields) + 1):
            sheet.cell(1, field, fields[field - 1][0])

        # 获取并写入数据段信息
        row = 2
        col = 1
        for row in range(2, len(results) + 2):
            for col in range(1, len(fields) + 1):
                sheet.cell(row, col, u'%s' % results[row - 2][col - 1])

    def save_to_excel_total(self, fileName, dic):
        # 创建数据库连接
        mysql = mysqlDB()
        conn = mysql.conn_db()
        cursor = mysql.cursor_db()
        # 创建Excel
        # excel=excel()
        workbook = self.create_excel()
        aaa = workbook.active  # 第一个sheet
        # 遍历dic 生成多个sheet
        for sql, sheet_name in dic.items():
            sql = sql
            sheet_name = sheet_name
            self.save_data1_to_excel(cursor, workbook, sql, sheet_name)

        # 删除默认第一个空白sheet
        workbook.remove(aaa)
        # 将数据存储到Excel中
        workbook.save(fileName)
        # 关闭数据库连接
        conn.commit()
        cursor.close()
        conn.close()


# 发送邮件
class sendmail(object):
    """docstring for sendmail"""

    def __init__(self, mailto_list, mailcc_list, mail_host, mail_user, mail_pass):
        self.mailto_list = mailto_list
        self.mailcc_list = mailcc_list
        self.mail_host = mail_host
        self.mail_user = mail_user
        self.mail_pass = mail_pass

# 参数：收件人，主题，正文，文件名集合（可发送多个文件），文件路径（文件在同一个路径）
    def send_mail(self, sub, content, files, path):
        mailto_list = self.mailto_list
        mail_host = self.mail_host
        mail_user = self.mail_user
        mail_pass = self.mail_pass
        me = mail_user + "<" + mail_user + ">"
        msg = MIMEMultipart()
        msg.attach(MIMEText(content, _subtype='html', _charset='utf-8'))
        msg['Subject'] = sub
        msg['From'] = me
        msg['To'] = ",".join(mailto_list)  # 将收件人列表以‘,’分隔
        msg['cc'] = ",".join(mailcc_list)  # 将抄送人列表以‘,’分割
        for file in files:
            if os.path.isfile(path + '/' + file):
                # 构造附件
                att = MIMEText(
                    open(path + '/' + file, 'rb').read(), 'base64', 'utf-8')
                att["Content-Type"] = 'application/octet-stream'
                att.add_header("Content-Disposition",
                               "attachment", filename=("gbk", "", file))
                msg.attach(att)
        try:
            #server = smtplib.SMTP()
            server = smtplib.SMTP_SSL(mail_host)
            # if mail_host == 'smtp.gmail.com':
            #     server.connect(mail_host, port=587)  # 连接服务器
            #     server.starttls()
            # else:
            #     server.connect(mail_host)
            server.login(mail_user, mail_pass)  # 登录操作
            server.sendmail(me, mailto_list, msg.as_string())
            server.close()
            print('Mail sent successfully')
            return True
        except Exception as e:
            print('Mail sent failed')
            print(sys.exc_info()[0], sys.exc_info()[1])
            return False


class seachFile(object):
    """查找path下文件名包含st字符的文件"""

    def __init__(self, path):
        self.path = path

    def fileList(self, st):
        filelist = []
        path = self.path
        files = glob(f'{path}/*{st}*', recursive=True)
        # files=glob.glob(path+r'/*'+st+r'*')   #python 2.7
        for file in files:
            filepath, filename = os.path.split(file)
        # for file in files:
            filelist.append(filename)
        return filelist


if __name__ == "__main__":
    sql1='''SQL1'''
    sql2='''sql2'''
	sql3='''sql3'''

    sheet_name1 = r"sheet_name1"
    sheet_name2 = r"sheet_name1"
    sheet_name3 = r"sheet_name1"

    dic = {sql1: sheet_name1, sql2: sheet_name2,
           sql3: sheet_name3}

    path = r"f:/1"  # 文件存放路径
    excelName = r'报表aaaa'
    datetime = datetime.datetime.now().strftime('%Y%m%d')
    fileName = path + r'/' + excelName + datetime + r'.xlsx'
    a = excel()
    a.save_to_excel_total(fileName, dic)

    # 邮件发送
    # date=time.strftime("%Y-%m-%d", time.localtime())
    # nowtime=datetime.datetime.now()
    # date_yes=((nowtime+datetime.timedelta(days=-1))).strftime('%Y%m%d')
    # date_week_ago=(nowtime+datetime.timedelta(days=-7)).strftime('%Y%m%d')

    mailto_list = ['lv-007@163.com', 'aaasgroup.cn']  # 收件人列表，以英文逗号分隔
    mailcc_list = []  # 抄送人列表，以英文逗号分割
    mail_host = "smtp.exmail.qq.com"  # 使用的邮箱的smtp服务器地址
    mail_user = "ssss@sygroup.cn"  # 用户名
    mail_pass = "pass# 授权码，不是密码，授权码要在邮箱设置中获取  我用的是密码
    path = r"f:/1"  # 文件目录
    sub = "测试test"  # 邮件标题
    content = "测试内容c"  # 邮件内容
    st = datetime    # 文件名包含的字符串
    files_a = seachFile(path)
    files = files_a.fileList(st)
    s = sendmail(mailto_list, mailcc_list, mail_host, mail_user, mail_pass)
    s.send_mail(sub, content, files, path)  
