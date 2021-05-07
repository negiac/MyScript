#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Date    : 2021-05-06 18:08:55
# @Author  : NegiaC (lv-007@163.com)



import pymysql.cursors
from openpyxl import Workbook

# 链接数据库的游标
connect = pymysql.Connect(
    host="192.168.23.27",
    port=3306,
    user='user',
    passwd='passwd',
    db='db',
    charset='utf8',
)
cursor = connect.cursor()

# 关闭数据库链接操作
# def clos_cursor():
#     return cursor.close();

# 读取数据库数据
def query_all(sql):
    cursor.execute(sql);
    fields = cursor.description
    results = cursor.fetchall();
    clos_cursor()
    return fields,results

#mysql 查询结果转成excel xlsx 格式
def mysql_results_to_xlsx(sheet_name,fields,results):
            workbook=Workbook()
            sheet = workbook.active
            sheet.title=sheet_name
            # 写上字段信息
            for field in range(0, len(fields)):
                field1=field+1
                sheet.cell(1, field1, fields[field][0])
            row = 2 #第一行已经被列名占据了
            col = 1
            for row in range(2, len(results) + 2):
                for col in range(1, len(fields)+1):
                    sheet.cell(row, col, u'%s' % results[row - 2][col-1])
            workbook.save(filename=sheet_name)

    # 关闭数据库链接操作
def clos_cursor():
    cursor.close();
    connect.close()

if __name__ == '__main__':
    sheet_name="test.xlsx"
    sql = "select * from table"
    fields,results=query_all(sql)
    mysql_results_to_xlsx(sheet_name,fields,results)
